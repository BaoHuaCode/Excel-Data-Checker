import openpyxl
from openpyxl.styles import Font
from pathlib import Path

def check_values(file_path):
    """Check the row value if none or < 0."""
    # Check if the file exist
    if not file_path.exists():
        print(f"Can't find file: {file_path}")
        return
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    # Loop the rows and get the values in each row.
    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            cell = sheet.cell(row=r, column=1)
            value = cell.value

        # Check if the value <0 or empty.
        is_invalid = False
        if value is None:
            is_invalid = True
        elif isinstance(value, (int,float)) and value <0:
            is_invalid = True

        if is_invalid:
            remark_cell = new_sheet.cell(row=r, column=3)
            remark_cell.value = 'Wrong Data'
            remark_cell.font = Font(bold=True, color="FF0000")
        else:
            remark_cell = new_sheet.cell(row=r, column=c)
            remark_cell.value = value
    # Save the file as a copy.
    save_path = file_path.parent/"correct_file_copy.xlsx"
    new_wb.save(save_path)
    print(f"Done, result save for {save_path}")
path = Path('/Users/qfhxsxzw/test/test.xlsx')
check_values(path)